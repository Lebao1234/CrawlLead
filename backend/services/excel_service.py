import io, csv, jwt, logging
from datetime import datetime
from flask import request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import SECRET_KEY

logger = logging.getLogger("CrawlLeadBackend")

def _safe_str(val):
    """Chuyển mọi kiểu dữ liệu về string an toàn cho Excel cell.
    Xử lý cả list ['userA', 'userB'] → 'userA, userB'
    """
    if val is None:
        return ""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val if v)
    return str(val)

EXPORT_HEADERS = [
    "Họ và Tên",
    "Chức vụ",
    "Công ty",
    "Email",
    "Số điện thoại",
    "Địa điểm",
    "LinkedIn URL",
    "Trạng thái",
    "Ngày thu thập",
    "Người thu thập"
]

FB_EXPORT_HEADERS = ["Author", "Group", "Content Snippet", "Link", "Date", "Crawled By"]
LK_EXPORT_HEADERS = ["Author", "Headline", "Content Snippet", "Type", "Reactions", "Link", "Date", "Crawled By"]

def _format_lead_row(lead):
    position = lead.get("position") or lead.get("title") or ""
    created_at_raw = lead.get("created_at") or ""
    created_at_nice = ""
    if created_at_raw:
        try:
            clean_dt = created_at_raw.split(".")[0]
            dt = datetime.fromisoformat(clean_dt)
            created_at_nice = dt.strftime("%d/%m/%Y %H:%M:%S")
        except Exception as err:
            logger.warning(f"Error parsing date '{created_at_raw}': {err}")
            created_at_nice = created_at_raw
    
    status_map = {
        "new": "Mới",
        "verified": "Đã xác minh",
        "contacted": "Đã liên hệ",
        "interested": "Quan tâm",
        "not_interested": "Không quan tâm",
        "duplicate": "Trùng lặp"
    }
    status_raw = lead.get("status") or "new"
    status_nice = status_map.get(status_raw, status_raw)
    
    return [
        _safe_str(lead.get("name")),
        _safe_str(position),
        _safe_str(lead.get("company")),
        _safe_str(lead.get("email")),
        _safe_str(lead.get("phone")),
        _safe_str(lead.get("location")),
        _safe_str(lead.get("linkedin_url")),
        status_nice,
        created_at_nice,
        _safe_str(lead.get("crawled_by"))  # list → "userA, userB"
    ]

def _format_fb_row(p):
    return [_safe_str(p.get("author")), _safe_str(p.get("group_name")), _safe_str(p.get("content_snippet")), _safe_str(p.get("post_url")), _safe_str(p.get("created_at")), _safe_str(p.get("crawled_by"))]

def _format_lk_row(p):
    return [_safe_str(p.get("author")), _safe_str(p.get("author_headline")), _safe_str(p.get("content_snippet")), _safe_str(p.get("post_type")), p.get("reactions_count") or 0, _safe_str(p.get("post_url")), _safe_str(p.get("created_at")), _safe_str(p.get("crawled_by"))]

def generic_export_csv(collection, headers, row_formatter, filename_prefix):
    token = request.args.get('token')
    if not token: return jsonify({'error': 'Token missing'}), 401
    try: 
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except Exception as err:
        logger.warning(f"Export CSV auth failed: {err}")
        return jsonify({'error': 'Invalid token'}), 401
    docs = list(collection.find({}, {"_id": 0}))
    if not docs: return jsonify({"error": "No data"}), 400
    output = io.StringIO()
    output.write("sep=,\r\n")
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(headers)
    for doc in docs: writer.writerow(row_formatter(doc))
    output.seek(0)
    file_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(file_bytes, mimetype="text/csv", as_attachment=True, download_name=filename)

def generic_export_xlsx(collection, headers, row_formatter, filename_prefix):
    token = request.args.get('token')
    if not token: return jsonify({'error': 'Token missing'}), 401
    try: 
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except Exception as err:
        logger.warning(f"Export XLSX auth failed: {err}")
        return jsonify({'error': 'Invalid token'}), 401
    docs = list(collection.find({}, {"_id": 0}))
    if not docs: return jsonify({"error": "No data"}), 400
    wb = Workbook()
    ws = wb.active
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, doc in enumerate(docs, 2):
        for col_idx, val in enumerate(row_formatter(doc), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(file_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

def parse_import_file(file_obj, filename, headers_map):
    ext = filename.split('.')[-1].lower()
    data = []
    if ext == 'csv':
        content = file_obj.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            doc = {}
            for k, v in row.items():
                if k in headers_map: doc[headers_map[k]] = v
            if doc: data.append(doc)
    elif ext == 'xlsx':
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            doc = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] in headers_map: doc[headers_map[headers[i]]] = val
            if doc: data.append(doc)
    return data
