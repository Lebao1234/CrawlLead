import io, csv, jwt, logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from pymongo import UpdateOne
from config import SECRET_KEY
from models.database import leads_collection
from services.auth_service import token_required
from services.dedup_service import _normalize_lead_item, _is_matching_lead
from services.excel_service import _format_lead_row, EXPORT_HEADERS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger("CrawlLeadBackend")
lead_bp = Blueprint('lead', __name__)

def get_all_leads_from_db():
    leads = list(leads_collection.find({}, {"_id": 0}))
    for l in leads:
        cb = l.get("crawled_by")
        if isinstance(cb, list):
            l["crawled_by"] = ", ".join(cb)
    return leads

def get_leads_with_id():
    return list(leads_collection.find())

def _get_filtered_leads(crawled_by=None):
    query = {}
    if crawled_by:
        query["crawled_by"] = crawled_by
    return list(leads_collection.find(query, {"_id": 0}))

@lead_bp.route("/api/leads", methods=["GET"])
@token_required
def get_leads(current_user):
    leads = get_all_leads_from_db()
    return jsonify({"leads": leads, "total": len(leads)})

@lead_bp.route("/api/leads", methods=["POST"])
@token_required
def add_leads(current_user):
    data = request.json
    new_leads = data if isinstance(data, list) else [data]
    username = current_user["username"]
    
    for l in new_leads:
        _normalize_lead_item(l)

    emails_to_check = [l["email_lower"] for l in new_leads if l.get("email_lower")]
    usernames_to_check = [l["lk_username"] for l in new_leads if l.get("lk_username")]
    keys_to_check = [l["name_comp_key"] for l in new_leads if l.get("name_comp_key")]

    or_conditions = []
    if emails_to_check:
        or_conditions.append({"email_lower": {"$in": emails_to_check}})
        or_conditions.append({"email": {"$in": emails_to_check}})
    if usernames_to_check:
        or_conditions.append({"lk_username": {"$in": usernames_to_check}})
    if keys_to_check:
        or_conditions.append({"name_comp_key": {"$in": keys_to_check}})

    existing_docs = list(leads_collection.find({"$or": or_conditions})) if or_conditions else []

    added, dupes = [], []
    now_iso = datetime.now().isoformat()
    bulk_updates = []
    to_insert = []

    for lead in new_leads:
        lead["created_at"] = now_iso
        
        match_doc = None
        for doc in existing_docs:
            if _is_matching_lead(lead, doc):
                match_doc = doc
                break

        if match_doc:
            update_fields = {}
            for k, v in lead.items():
                if v and v != "Chưa có" and k not in ["status", "crawled_by", "created_at", "_id"]:
                    update_fields[k] = v

            cb = match_doc.get("crawled_by")
            cb_list = [cb] if isinstance(cb, str) else (cb.copy() if isinstance(cb, list) else [])
            if username not in cb_list:
                cb_list.append(username)

            update_fields["crawled_by"] = cb_list
            match_doc.update(update_fields)

            if "_id" in match_doc:
                bulk_updates.append(UpdateOne({"_id": match_doc["_id"]}, {"$set": update_fields}))

            res_lead = match_doc.copy()
            res_lead.pop("_id", None)
            if isinstance(res_lead.get("crawled_by"), list):
                res_lead["crawled_by"] = ", ".join(res_lead["crawled_by"])
            dupes.append(res_lead)
        else:
            lead["status"] = "new"
            lead["crawled_by"] = [username]
            to_insert.append(lead)
            
            res_lead = lead.copy()
            res_lead.pop("_id", None)
            res_lead["crawled_by"] = username
            added.append(res_lead)

    if bulk_updates:
        leads_collection.bulk_write(bulk_updates, ordered=False)

    if to_insert:
        leads_collection.insert_many(to_insert)

    logger.info(f"User {username} added {len(added)} new leads ({len(dupes)} duplicates merged)")
    return jsonify({"added": len(added), "duplicates": len(dupes), "leads": added})

@lead_bp.route("/api/leads/<int:idx>", methods=["GET", "DELETE", "POST"])
@lead_bp.route("/api/leads/<int:idx>/delete", methods=["GET", "DELETE", "POST"])
@token_required
def delete_lead(current_user, idx):
    leads = get_leads_with_id()
    if idx < 0 or idx >= len(leads):
        return jsonify({"error": "Not found"}), 404
    
    doc = leads[idx]
    cb = doc.get("crawled_by")
    username = current_user["username"]

    if isinstance(cb, list):
        if username in cb and len(cb) > 1:
            leads_collection.update_one({"_id": doc["_id"]}, {"$pull": {"crawled_by": username}})
        else:
            leads_collection.delete_one({"_id": doc["_id"]})
    else:
        leads_collection.delete_one({"_id": doc["_id"]})
        
    return jsonify({"ok": True, "deleted": 1})

@lead_bp.route("/api/leads/delete", methods=["GET", "DELETE", "POST"])
@token_required
def delete_lead_query(current_user):
    idx_str = request.args.get('idx')
    if idx_str is None and request.is_json and request.json:
        idx_str = request.json.get('idx')
    if idx_str is None:
        return jsonify({"error": "Missing idx parameter"}), 400
    try:
        idx = int(idx_str)
        return delete_lead(current_user, idx)
    except ValueError:
        return jsonify({"error": "Invalid idx parameter"}), 400

@lead_bp.route("/api/leads/bulk-delete", methods=["GET", "POST", "DELETE"])
@token_required
def bulk_delete_leads(current_user):
    data = request.json or {}
    if not data and request.args.get("indices"):
        data = {"indices": [int(x) for x in request.args.get("indices").split(",") if x.isdigit()]}
    indices = set(data.get("indices", []))
    leads = get_leads_with_id()
    username = current_user["username"]
    
    selected_docs = [leads[i] for i in range(len(leads)) if i in indices]
    deleted_count = 0

    for doc in selected_docs:
        cb = doc.get("crawled_by")
        if isinstance(cb, list):
            if username in cb and len(cb) > 1:
                leads_collection.update_one({"_id": doc["_id"]}, {"$pull": {"crawled_by": username}})
            else:
                leads_collection.delete_one({"_id": doc["_id"]})
        else:
            leads_collection.delete_one({"_id": doc["_id"]})
        deleted_count += 1
            
    return jsonify({"ok": True, "deleted": deleted_count})

@lead_bp.route("/api/leads/clear", methods=["GET", "POST", "DELETE"])
@token_required
def clear_leads(current_user):
    username = current_user["username"]
    
    try:
        leads_collection.update_many({"crawled_by": {"$type": "array"}}, {"$pull": {"crawled_by": username}})
    except Exception as e:
        logger.warning(f"Error pulling from array crawled_by: {e}")

    res = leads_collection.delete_many({
        "$or": [
            {"crawled_by": username},
            {"crawled_by": []},
            {"crawled_by": ""},
            {"crawled_by": None},
            {"crawled_by": {"$exists": False}}
        ]
    })
    
    deleted = res.deleted_count
    if deleted == 0:
        res_all = leads_collection.delete_many({})
        deleted = res_all.deleted_count
        
    return jsonify({"ok": True, "deleted": deleted})

@lead_bp.route("/api/leads/<int:idx>/verify", methods=["POST"])
@token_required
def verify_lead(current_user, idx):
    leads = get_leads_with_id()
    if idx < 0 or idx >= len(leads):
        return jsonify({"error": "Not found"}), 404
    
    doc_to_verify = leads[idx]
    leads_collection.update_one({"_id": doc_to_verify["_id"]}, {"$set": {"status": "verified"}})
    return jsonify({"ok": True})

@lead_bp.route("/api/export/csv", methods=["GET"])
def export_csv():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Token is missing'}), 401
    try:
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except Exception as err:
        logger.warning(f"Export CSV token validation failed: {err}")
        return jsonify({'error': 'Token is invalid'}), 401

    crawled_by = request.args.get('crawled_by', '').strip()
    leads = _get_filtered_leads(crawled_by if crawled_by else None)
    if not leads:
        return jsonify({"error": "No leads"}), 400
    
    output = io.StringIO()
    output.write("sep=,\r\n")
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(EXPORT_HEADERS)
    
    for lead in leads:
        writer.writerow(_format_lead_row(lead))
    
    output.seek(0)
    file_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    
    suffix = f"_{crawled_by}" if crawled_by else ""
    filename = f"leads{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        file_bytes,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )

@lead_bp.route("/api/export/xlsx", methods=["GET"])
def export_xlsx():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Token is missing'}), 401
    try:
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except Exception as err:
        logger.warning(f"Export XLSX token validation failed: {err}")
        return jsonify({'error': 'Token is invalid'}), 401

    crawled_by = request.args.get('crawled_by', '').strip()
    leads = _get_filtered_leads(crawled_by if crawled_by else None)
    if not leads:
        return jsonify({"error": "No leads"}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=False)
    even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    for row_idx, lead in enumerate(leads, 2):
        row_data = _format_lead_row(lead)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill
    
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        max_len = len(header)
        for row_idx in range(2, len(leads) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(cell_val) > max_len:
                max_len = len(cell_val)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)
    
    ws.freeze_panes = "A2"
    
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    
    suffix = f"_{crawled_by}" if crawled_by else ""
    filename = f"leads{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        file_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
