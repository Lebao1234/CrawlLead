from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import json, csv, os, io, re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import mimetypes
from pymongo import MongoClient
from dotenv import load_dotenv
import jwt
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

# Đọc file .env ở thư mục gốc
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

# Đảm bảo hệ thống nhận diện đúng định dạng file CSS và JS
mimetypes.add_type('text/css', '.css')
mimetypes.add_type('application/javascript', '.js')

base_dir = os.path.dirname(os.path.abspath(__file__))
web_app_dir = os.path.join(base_dir, "..", "frontend")

app = Flask(__name__, static_folder=web_app_dir, static_url_path="/static_assets_hidden")
CORS(app)

# ==========================================
# KHỞI TẠO KẾT NỐI MONGODB
# ==========================================
MONGO_URL = os.environ.get("MongoURL")
if not MONGO_URL:
    raise ValueError("Missing MongoURL in .env file")

client = MongoClient(MONGO_URL)
db = client["CrawlLead"]
leads_collection = db["Linkedin"]
fb_collection = db["Facebook"]
lk_posts_collection = db["LinkedInPosts"]
users_collection = db["Users"]

SECRET_KEY = os.environ.get("SECRET_KEY", "leadfinder_secret_key_123")

# ==========================================
# PHẦN 0: XÁC THỰC (AUTHENTICATION)
# ==========================================

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            if auth_header.startswith('Bearer '):
                token = auth_header.split(" ")[1]
        
        if not token:
            return jsonify({'error': 'Token is missing!'}), 401
        
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            current_user = users_collection.find_one({"username": data['username']})
            if not current_user:
                raise Exception("User not found")
        except:
            return jsonify({'error': 'Token is invalid or expired!'}), 401
            
        return f(current_user, *args, **kwargs)
    return decorated

@app.route("/api/register", methods=["POST"])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Missing username or password'}), 400
        
    if users_collection.find_one({"username": username}):
        return jsonify({'error': 'User already exists'}), 400
        
    hashed_password = generate_password_hash(password)
    users_collection.insert_one({"username": username, "password": hashed_password})
    return jsonify({'ok': True, 'message': 'Registered successfully'})

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    user = users_collection.find_one({"username": username})
    if not user or not check_password_hash(user['password'], password):
        return jsonify({'error': 'Invalid credentials'}), 401
    
    token = jwt.encode({
        'username': user['username'],
        'exp': datetime.utcnow() + timedelta(days=30)
    }, SECRET_KEY, algorithm="HS256")
    
    return jsonify({'token': token, 'username': user['username']})

# ==========================================
# PHẦN 1: PHỤC VỤ GIAO DIỆN WEB (FRONTEND)
# ==========================================

@app.route("/<path:filename>")
def serve_static(filename):
    file_path = os.path.join(web_app_dir, filename)
    if os.path.exists(file_path):
        mimetype = None
        if filename.endswith(".css"):
            mimetype = "text/css"
        elif filename.endswith(".js"):
            mimetype = "application/javascript"
        return send_file(file_path, mimetype=mimetype)
    return jsonify({"error": "Not found"}), 404

@app.route("/")
def index():
    return send_file(os.path.join(web_app_dir, "index.html"))

# ==========================================
# PHẦN 2: CÁC API ENDPOINTS XỬ LÝ DỮ LIỆU TỪ MONGODB
# ==========================================

def normalize_linkedin_url(url):
    if not url:
        return ""
    # Standardize to lowercase and strip whitespace
    url = url.strip().split("?")[0]
    if url.endswith("/"):
        url = url[:-1]
    # Standardize subdomains
    url = re.sub(r'^https?://[a-z]{2,3}\.linkedin\.com', 'https://www.linkedin.com', url)
    url = re.sub(r'^https?://linkedin\.com', 'https://www.linkedin.com', url)
    return url

def get_all_leads_from_db():
    return list(leads_collection.find({}, {"_id": 0}))

def get_leads_with_id():
    return list(leads_collection.find())

@app.route("/api/leads", methods=["GET"])
@token_required
def get_leads(current_user):
    leads = get_all_leads_from_db()
    return jsonify({"leads": leads, "total": len(leads)})

@app.route("/api/leads", methods=["POST"])
@token_required
def add_leads(current_user):
    data = request.json
    new_leads = data if isinstance(data, list) else [data]
    
    added, dupes = [], []
    
    for lead in new_leads:
        lead.pop("_id", None)
        lead["created_at"] = datetime.now().isoformat()
        lead["crawled_by"] = current_user["username"] # Lưu người thu thập
        
        query = {"$or": []}
        if lead.get("email") and lead["email"] != "Chưa có" and lead["email"] != "":
            query["$or"].append({"email": {"$regex": f"^{lead['email']}$", "$options": "i"}})
        
        if lead.get("linkedin_url"):
            norm_url = normalize_linkedin_url(lead["linkedin_url"])
            lead["linkedin_url"] = norm_url
            
            # Trích xuất username để tìm kiếm trùng lặp cả trong dữ liệu cũ chưa chuẩn hóa
            if "/in/" in norm_url:
                username = norm_url.split("/in/")[-1].split("/")[0]
                if username:
                    pattern = f"linkedin\\.com/in/{re.escape(username)}/?(?:$|\\?)"
                    query["$or"].append({"linkedin_url": {"$regex": pattern, "$options": "i"}})
                else:
                    query["$or"].append({"linkedin_url": norm_url})
            else:
                query["$or"].append({"linkedin_url": norm_url})
            
        if not query["$or"]:
            existing_lead = None
        else:
            existing_lead = leads_collection.find_one(query)
            
        if existing_lead:
            update_data = {}
            for k, v in lead.items():
                if v and v != "Chưa có" and k != "status":
                    update_data[k] = v
            
            # Đảm bảo lưu URL đã chuẩn hóa
            if lead.get("linkedin_url"):
                update_data["linkedin_url"] = lead["linkedin_url"]
                
            if update_data:
                leads_collection.update_one({"_id": existing_lead["_id"]}, {"$set": update_data})
                
            existing_lead.pop("_id", None)
            existing_lead.update(update_data)
            dupes.append(existing_lead)
        else:
            lead["status"] = "new"
            leads_collection.insert_one(lead)
            inserted_lead = lead.copy()
            inserted_lead.pop("_id", None)
            added.append(inserted_lead)

    return jsonify({"added": len(added), "duplicates": len(dupes), "leads": added})

@app.route("/api/leads/<int:idx>", methods=["DELETE"])
@token_required
def delete_lead(current_user, idx):
    leads = get_leads_with_id()
    if idx < 0 or idx >= len(leads):
        return jsonify({"error": "Not found"}), 404
    
    doc_to_delete = leads[idx]
    leads_collection.delete_one({"_id": doc_to_delete["_id"]})
    return jsonify({"ok": True})

@app.route("/api/leads/bulk-delete", methods=["POST"])
@token_required
def bulk_delete_leads(current_user):
    data = request.json
    indices = set(data.get("indices", []))
    leads = get_leads_with_id()
    
    ids_to_delete = [leads[i]["_id"] for i in range(len(leads)) if i in indices]
    
    if ids_to_delete:
        leads_collection.delete_many({"_id": {"$in": ids_to_delete}})
        
    return jsonify({"ok": True, "deleted": len(ids_to_delete)})

@app.route("/api/leads/clear", methods=["POST"])
@token_required
def clear_leads(current_user):
    leads_collection.delete_many({})
    return jsonify({"ok": True})

@app.route("/api/leads/<int:idx>/verify", methods=["POST"])
@token_required
def verify_lead(current_user, idx):
    leads = get_leads_with_id()
    if idx < 0 or idx >= len(leads):
        return jsonify({"error": "Not found"}), 404
    
    doc_to_verify = leads[idx]
    leads_collection.update_one({"_id": doc_to_verify["_id"]}, {"$set": {"status": "verified"}})
    return jsonify({"ok": True})

@app.route("/api/stats", methods=["GET"])
def stats():
    # Public API or optional auth? Let's make it public so checkBackend works easily without token
    total = leads_collection.count_documents({})
    verified = leads_collection.count_documents({"status": "verified"})
    dupes = leads_collection.count_documents({"status": "duplicate"})
    new = leads_collection.count_documents({"status": "new"})
    lk_posts = lk_posts_collection.count_documents({})
    fb_posts = fb_collection.count_documents({})
    
    return jsonify({"total": total, "verified": verified, "duplicates": dupes, "new": new, "lk_posts": lk_posts, "fb_posts": fb_posts})

@app.route("/api/crawlers", methods=["GET"])
def get_crawlers():
    """Trả về danh sách tất cả người thu thập (crawled_by) duy nhất từ collection Linkedin"""
    crawlers = leads_collection.distinct("crawled_by")
    # Lọc bỏ giá trị None/rỗng
    crawlers = [c for c in crawlers if c]
    crawlers.sort()
    return jsonify({"crawlers": crawlers})

def _format_lead_row(lead):
    """Helper: format một lead thành row data cho export CSV/XLSX"""
    position = lead.get("position") or lead.get("title") or ""
    
    created_at_raw = lead.get("created_at") or ""
    created_at_nice = ""
    if created_at_raw:
        try:
            clean_dt = created_at_raw.split(".")[0]
            dt = datetime.fromisoformat(clean_dt)
            created_at_nice = dt.strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            created_at_nice = created_at_raw
    
    status_map = {
        "new": "Mới",
        "verified": "Đã xác minh",
        "contacted": "Đã liên hệ",
        "interested": "Quan tâm",
        "not_interested": "Không quan tâm",
        "duplicate": "Trùng lặp"
    }
    status_raw = lead.get("status") or "new"
    status_nice = status_map.get(status_raw, status_raw)
    
    return [
        lead.get("name") or "",
        position,
        lead.get("company") or "",
        lead.get("email") or "",
        lead.get("phone") or "",
        lead.get("location") or "",
        lead.get("linkedin_url") or "",
        status_nice,
        created_at_nice,
        lead.get("crawled_by") or ""
    ]

EXPORT_HEADERS = [
    "Họ và Tên",
    "Chức vụ",
    "Công ty",
    "Email",
    "Số điện thoại",
    "Địa điểm",
    "LinkedIn URL",
    "Trạng thái",
    "Ngày thu thập",
    "Người thu thập"
]

def _get_filtered_leads(crawled_by=None):
    """Lấy leads từ DB, lọc theo crawled_by nếu có"""
    query = {}
    if crawled_by:
        query["crawled_by"] = crawled_by
    return list(leads_collection.find(query, {"_id": 0}))

@app.route("/api/export/csv", methods=["GET"])
def export_csv():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Token is missing'}), 401
    try:
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except:
        return jsonify({'error': 'Token is invalid'}), 401

    crawled_by = request.args.get('crawled_by', '').strip()
    leads = _get_filtered_leads(crawled_by if crawled_by else None)
    if not leads:
        return jsonify({"error": "No leads"}), 400
    
    output = io.StringIO()
    output.write("sep=,\r\n")
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(EXPORT_HEADERS)
    
    for lead in leads:
        writer.writerow(_format_lead_row(lead))
    
    output.seek(0)
    file_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    
    suffix = f"_{crawled_by}" if crawled_by else ""
    filename = f"leads{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        file_bytes,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )

@app.route("/api/export/xlsx", methods=["GET"])
def export_xlsx():
    token = request.args.get('token')
    if not token:
        return jsonify({'error': 'Token is missing'}), 401
    try:
        jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except:
        return jsonify({'error': 'Token is invalid'}), 401

    crawled_by = request.args.get('crawled_by', '').strip()
    leads = _get_filtered_leads(crawled_by if crawled_by else None)
    if not leads:
        return jsonify({"error": "No leads"}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Style cho header
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    
    # Ghi header
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Style cho data rows
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=False)
    even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Ghi data
    for row_idx, lead in enumerate(leads, 2):
        row_data = _format_lead_row(lead)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill
    
    # Auto-fit cột (ước lượng chiều rộng)
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        max_len = len(header)
        for row_idx in range(2, len(leads) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(cell_val) > max_len:
                max_len = len(cell_val)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Lưu vào buffer
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    
    suffix = f"_{crawled_by}" if crawled_by else ""
    filename = f"leads{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        file_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# ==========================================
# PHẦN 3: CÁC API ENDPOINTS XỬ LÝ DỮ LIỆU TỪ FACEBOOK
# ==========================================

def get_all_fb_from_db():
    return list(fb_collection.find({}, {"_id": 0}))

def get_fb_with_id():
    return list(fb_collection.find())

@app.route("/api/facebook", methods=["GET"])
@token_required
def get_fb_posts(current_user):
    posts = get_all_fb_from_db()
    return jsonify({"posts": posts, "total": len(posts)})

@app.route("/api/facebook", methods=["POST"])
@token_required
def add_fb_posts(current_user):
    data = request.json
    new_posts = data if isinstance(data, list) else [data]
    
    added, dupes = [], []
    
    for post in new_posts:
        post.pop("_id", None)
        post["created_at"] = datetime.now().isoformat()
        post["crawled_by"] = current_user["username"] # Lưu người thu thập
        
        query = None
        if post.get("post_url"):
            query = {"post_url": post["post_url"]}
        
        existing_post = fb_collection.find_one(query) if query else None
            
        if existing_post:
            update_data = {}
            for k, v in post.items():
                if v and k != "status":
                    update_data[k] = v
            
            if update_data:
                fb_collection.update_one({"_id": existing_post["_id"]}, {"$set": update_data})
                
            existing_post.pop("_id", None)
            existing_post.update(update_data)
            dupes.append(existing_post)
        else:
            fb_collection.insert_one(post)
            inserted_post = post.copy()
            inserted_post.pop("_id", None)
            added.append(inserted_post)

    return jsonify({"added": len(added), "duplicates": len(dupes), "posts": added})

@app.route("/api/facebook/<int:idx>", methods=["DELETE"])
@token_required
def delete_fb_post(current_user, idx):
    posts = get_fb_with_id()
    if idx < 0 or idx >= len(posts):
        return jsonify({"error": "Not found"}), 404
    
    doc_to_delete = posts[idx]
    fb_collection.delete_one({"_id": doc_to_delete["_id"]})
    return jsonify({"ok": True})

@app.route("/api/facebook/clear", methods=["POST"])
@token_required
def clear_fb_posts(current_user):
    fb_collection.delete_many({})
    return jsonify({"ok": True})

# ==========================================
# PHẦN 4: CÁC API ENDPOINTS XỬ LÝ LINKEDIN POSTS
# ==========================================

def get_all_lk_posts_from_db():
    return list(lk_posts_collection.find({}, {"_id": 0}))

def get_lk_posts_with_id():
    return list(lk_posts_collection.find())

@app.route("/api/lk-posts", methods=["GET"])
@token_required
def get_lk_posts(current_user):
    posts = get_all_lk_posts_from_db()
    return jsonify({"posts": posts, "total": len(posts)})

@app.route("/api/lk-posts", methods=["POST"])
@token_required
def add_lk_posts(current_user):
    data = request.json
    new_posts = data if isinstance(data, list) else [data]
    added, dupes = [], []
    for post in new_posts:
        post.pop("_id", None)
        post["created_at"] = datetime.now().isoformat()
        post["crawled_by"] = current_user["username"]
        query = None
        if post.get("post_url"):
            query = {"post_url": post["post_url"]}
        existing_post = lk_posts_collection.find_one(query) if query else None
        if existing_post:
            update_data = {}
            for k, v in post.items():
                if v and k != "status":
                    update_data[k] = v
            if update_data:
                lk_posts_collection.update_one({"_id": existing_post["_id"]}, {"$set": update_data})
            existing_post.pop("_id", None)
            existing_post.update(update_data)
            dupes.append(existing_post)
        else:
            lk_posts_collection.insert_one(post)
            inserted_post = post.copy()
            inserted_post.pop("_id", None)
            added.append(inserted_post)
    return jsonify({"added": len(added), "duplicates": len(dupes), "posts": added})

@app.route("/api/lk-posts/<int:idx>", methods=["DELETE"])
@token_required
def delete_lk_post(current_user, idx):
    posts = get_lk_posts_with_id()
    if idx < 0 or idx >= len(posts):
        return jsonify({"error": "Not found"}), 404
    doc_to_delete = posts[idx]
    lk_posts_collection.delete_one({"_id": doc_to_delete["_id"]})
    return jsonify({"ok": True})

@app.route("/api/lk-posts/bulk-delete", methods=["POST"])
@token_required
def bulk_delete_lk_posts(current_user):
    data = request.json
    indices = set(data.get("indices", []))
    posts = get_lk_posts_with_id()
    ids_to_delete = [posts[i]["_id"] for i in range(len(posts)) if i in indices]
    if ids_to_delete:
        lk_posts_collection.delete_many({"_id": {"$in": ids_to_delete}})
    return jsonify({"ok": True, "deleted": len(ids_to_delete)})

@app.route("/api/lk-posts/clear", methods=["POST"])
@token_required
def clear_lk_posts(current_user):
    lk_posts_collection.delete_many({})
    return jsonify({"ok": True})


# ==========================================
# PHẦN 5: EXPORT & IMPORT FB VÀ LK POSTS
# ==========================================

FB_EXPORT_HEADERS = ["Author", "Group", "Content Snippet", "Link", "Date", "Crawled By"]
def _format_fb_row(p):
    return [p.get("author") or "", p.get("group_name") or "", p.get("content_snippet") or "", p.get("post_url") or "", p.get("created_at") or "", p.get("crawled_by") or ""]

LK_EXPORT_HEADERS = ["Author", "Headline", "Content Snippet", "Type", "Reactions", "Link", "Date", "Crawled By"]
def _format_lk_row(p):
    return [p.get("author") or "", p.get("author_headline") or "", p.get("content_snippet") or "", p.get("post_type") or "", p.get("reactions_count") or 0, p.get("post_url") or "", p.get("created_at") or "", p.get("crawled_by") or ""]

def generic_export_csv(collection, headers, row_formatter, filename_prefix):
    token = request.args.get('token')
    if not token: return jsonify({'error': 'Token missing'}), 401
    try: jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except: return jsonify({'error': 'Invalid token'}), 401
    docs = list(collection.find({}, {"_id": 0}))
    if not docs: return jsonify({"error": "No data"}), 400
    output = io.StringIO()
    output.write("sep=,\r\n")
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(headers)
    for doc in docs: writer.writerow(row_formatter(doc))
    output.seek(0)
    file_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(file_bytes, mimetype="text/csv", as_attachment=True, download_name=filename)

def generic_export_xlsx(collection, headers, row_formatter, filename_prefix):
    token = request.args.get('token')
    if not token: return jsonify({'error': 'Token missing'}), 401
    try: jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except: return jsonify({'error': 'Invalid token'}), 401
    docs = list(collection.find({}, {"_id": 0}))
    if not docs: return jsonify({"error": "No data"}), 400
    wb = Workbook()
    ws = wb.active
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, doc in enumerate(docs, 2):
        for col_idx, val in enumerate(row_formatter(doc), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(file_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

@app.route("/api/facebook/export/csv", methods=["GET"])
def export_fb_csv(): return generic_export_csv(fb_collection, FB_EXPORT_HEADERS, _format_fb_row, "fb_posts")

@app.route("/api/facebook/export/xlsx", methods=["GET"])
def export_fb_xlsx(): return generic_export_xlsx(fb_collection, FB_EXPORT_HEADERS, _format_fb_row, "fb_posts")

@app.route("/api/lk-posts/export/csv", methods=["GET"])
def export_lk_csv(): return generic_export_csv(lk_posts_collection, LK_EXPORT_HEADERS, _format_lk_row, "lk_posts")

@app.route("/api/lk-posts/export/xlsx", methods=["GET"])
def export_lk_xlsx(): return generic_export_xlsx(lk_posts_collection, LK_EXPORT_HEADERS, _format_lk_row, "lk_posts")

def parse_import_file(file_obj, filename, headers_map):
    ext = filename.split('.')[-1].lower()
    data = []
    if ext == 'csv':
        content = file_obj.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            doc = {}
            for k, v in row.items():
                if k in headers_map: doc[headers_map[k]] = v
            if doc: data.append(doc)
    elif ext == 'xlsx':
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            doc = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] in headers_map: doc[headers_map[headers[i]]] = val
            if doc: data.append(doc)
    return data

@app.route("/api/facebook/import", methods=["POST"])
@token_required
def import_fb(current_user):
    if 'file' not in request.files: return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    if not file.filename.endswith(('.csv', '.xlsx')): return jsonify({"error": "Invalid file type"}), 400
    try:
        data = parse_import_file(file, file.filename, {v: k for k, v in zip(FB_EXPORT_HEADERS, ["author", "group_name", "content_snippet", "post_url", "created_at", "crawled_by"])})
        added, dupes = 0, 0
        for post in data:
            post["crawled_by"] = current_user["username"]
            if not post.get("created_at"): post["created_at"] = datetime.now().isoformat()
            query = {"post_url": post["post_url"]} if post.get("post_url") else None
            existing = fb_collection.find_one(query) if query else None
            if existing: dupes += 1
            else:
                fb_collection.insert_one(post)
                added += 1
        return jsonify({"added": added, "duplicates": dupes})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/lk-posts/import", methods=["POST"])
@token_required
def import_lk(current_user):
    if 'file' not in request.files: return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    if not file.filename.endswith(('.csv', '.xlsx')): return jsonify({"error": "Invalid file type"}), 400
    try:
        data = parse_import_file(file, file.filename, {v: k for k, v in zip(LK_EXPORT_HEADERS, ["author", "author_headline", "content_snippet", "post_type", "reactions_count", "post_url", "created_at", "crawled_by"])})
        added, dupes = 0, 0
        for post in data:
            post["crawled_by"] = current_user["username"]
            if not post.get("created_at"): post["created_at"] = datetime.now().isoformat()
            query = {"post_url": post["post_url"]} if post.get("post_url") else None
            existing = lk_posts_collection.find_one(query) if query else None
            if existing: dupes += 1
            else:
                lk_posts_collection.insert_one(post)
                added += 1
        return jsonify({"added": added, "duplicates": dupes})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)
